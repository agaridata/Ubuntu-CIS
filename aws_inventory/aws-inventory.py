import boto3
import json
import pprint
import getopt
import sys
import os
import time
import datetime
import yaml
import pyjq
import re
import hashlib
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Fill, Alignment, Color, PatternFill
from openpyxl.styles import colors as Colors
from ruamel.yaml import YAML
from ruamel.yaml.compat import StringIO

pp = pprint.PrettyPrinter(indent=1)

class MyYAML(YAML):
    def dump(self, data, stream=None, **kw):
        inefficient = False
        if stream is None:
            inefficient = True
            stream = StringIO()
        YAML.dump(self, data, stream, **kw)
        if inefficient:
            return stream.getvalue()

def conf():
    rmap = {}
    region = 'us-east-1'
    filename = ''
    of = ''
    debug = 0
    resource = 'all'
    options, remainder = getopt.getopt(sys.argv[1:], 'f:r:d:o:R:h', ['filename=', 
                                                         'region=', 
                                                         'debug=', 
                                                         'output=', 
                                                         'resource=', 
                                                         'help',
                                                         ])
    for opt, arg in options:
        if opt in ('-f', '--filename'):
            filename = arg
        elif opt in ('-r', '--region'):
            region = arg
        elif opt in ('-d', '--debug'):
            debug = arg
        elif opt in ('-o', '--output'):
            of = arg
        elif opt in ('-R', '--resource'):
            resource = arg
        elif opt in ('-h', '--help'):
            help_ = True

    if debug > 9 or debug < 0:
        errx("debug has incorrect value")

    if not filename:
        errx("filename not defined")
    if not of:
        errx("output filename not defined")

    with open(filename) as this:
        rmap = yaml.load(this)
    rmap['region'] = region
    rmap['debug'] = debug
    rmap['of'] = of
    return rmap

def help_():
    Progname = os.path.basename(__file__)
    print
    print("Usage:")
    print("python %s --filename <FILENAME> -r <REGION>" % Progname)
    print("       --filename - filename with resources specification in YAML")
    print("       --region - aws-region, default: us-east-1")
    print("       --debug - verbose output(0..9), default: 0")
    print("       --output - write to file (.xlsx)")
    print
    print("Example:")
    print("python %s -f resources.yaml -r us-east-1 -o example.xlsx" % Progname)
    print
    sys.exit('')

def errx(msg):
    print
    print("[ERROR]: %s" % msg)
    help_()


class ClassHolder(object):
    def __init__(self):
        self.classes = {}

    def add_class(self, c):
        self.classes[c.__name__] = c

    def __getitem__(self, n):
        return self.classes[n]

class aws(object):
    """Parent boto3 class"""
    name = "AWS"

    MaxItems = 100

    def __init__(self, conf=None, region='us-east-1'):
        """Constructor"""
        self.conf = [] if conf is None else conf
        self.region = region
        self.client = NotImplemented
        self.name = self.getname()
        self.attributes = conf[self.name]
        self.description = self.attributes['description'] if 'description' in self.attributes else self.name
        self.client = boto3.client(self.service, region_name=self.region)
        self.cfdata = self.init_cellformat_data()
        self.colors = self.rebase_colors()

    def init_cellformat_data(self):
        key = 'CellFormat'
        data = self.attributes[key] if key in self.attributes else {}
        colors = {}
        c = 'colors'
        for item in data:
            if c in data[item]:
                for k,v in data[item][c].iteritems():
                    for pattern in v:
                        colors[pattern] = k
                data[item][c] = colors
        return data
              
    def paginator(self, call=None, key=None, **kwargs):
        a = call(**kwargs)
        base = a
        while 'NextToken' in a:
            a = call(NextToken=a['NextToken'], **kwargs)
            base[key].extend(a[key])
        return base

    def boto_method_handler(self, call=None, **kwargs):
        data = None
        if call is None or kwargs is None:
            return data
        try:
            data = call(**kwargs)
        except:
            return None

        if 'ResponseMetadata' in data:
            del data['ResponseMetadata']
        return data

    def describe(self):
        return []
        
    def getname(self):
        return self.__class__.__name__

    def jqfilter(self):
        a = self.describe()
        if a is None:
            return []
        try:
            return pyjq.all(self.attributes['filter'], self.json_datetime_convert(a))
        except:
            return []

    def rebase_colors(self):
        colors = {
            False: 'EFAA9C',
            True: 'CDEACE'
        }

        data = self.attributes['colors'] if 'colors' in self.attributes else {}
        for k,v in data.iteritems():
            for i in v:
                colors[i] = k
        return colors

    def json_to_string(self, iterkeys=[], basekey=None, data=None, rc=1, header='', out=''):
        delim = ' - '
        vdata = None
        colors = self.colors
        if data is None or iterkeys == []:
            return (out, rc)
        if basekey is None:
            vdata = data
        elif basekey in data:
            vdata = data[basekey]
        else:
            return (out, rc)
        for item in vdata:
            out += '\n' if len(out) > 0 else ''
            out += str(rc) + '. ' + header
            irc = 0
            for i in iterkeys:
                if i in item:
                    out += delim if irc > 0 else ''
                    out += item[i]
                    irc += 1
            rc += 1
        return (out, rc)

    """
    Functions to serialize datetime.datetime in boto3 data
    https://stackoverflow.com/questions/11875770/how-to-overcome-datetime-datetime-not-json-serializable
    """

    def json_datetime_serialize(self, o):
        if isinstance(o, (datetime.date, datetime.datetime)):
            return o.isoformat()

    def json_datetime_convert(self, data):
        c = json.dumps(data, default=self.json_datetime_serialize)
        return json.loads(c)

    def format_cell_yaml(self, value=None):
        yaml = MyYAML()
        yaml.preserve_quotes = True
        records = json.loads(value)
        if records is None: 
            return False
        if (type(records) is list and records == []):
            return False
        if (type(records) is dict and records == {}):
            return False
        return yaml.dump(records)

    def format_cell_loop(self, value=None, field=None, header='', rc=1, out=''):
        data = self.cfdata
        ikey = 'iterate_by'
        cf = data[field][ikey] if field in data and ikey in data[field] else None
        if cf is None:
            out = self.format_cell_yaml(value=value)
            return (out, rc)
        
        records = json.loads(value)
        for item in cf:
            for k,v in item.iteritems():
                (out, rc) = self.json_to_string(iterkeys=v, basekey=k, data=records, rc=rc, header=header, out=out)
        return (out, rc)

    def format_cell(self, value=None, field=None):
        return value

    def set_cell_color(self, value=None, field=None):
        color = None
        if value is None:
            return color
        colors = self.colors
        if value in colors:
            color = Colors.Color(rgb=colors[value])
        return color

    def xlsx(self, wb, highlight):

        a = self.jqfilter()
        colors = self.colors
        ws = wb.create_sheet(self.name)
        maxwidth = 100
        row = 1
        # header

        for i in a:
            column = 1
            for k in i:
                c = ws.cell(row=row, column=column, value=k)
                c.style = highlight
                column += 1
            break

        # Content

        row += 1
        for i in a:
            column = 1
            for k in i:
                if k in self.cfdata:
                    value = self.format_cell(value=i[k], field=k)
                else:
                    value = i[k]
                    if value is None or value == 'none':
                        value = False
                c = ws.cell(row=row, column=column, value=value)
                color = self.set_cell_color(value=value, field=k)
                if color is not None:
                    ws.cell(row=row, column=column).fill = PatternFill(patternType='solid', fgColor=color)
                ws.cell(row=row, column=column).alignment = Alignment(vertical='top', wrap_text=True, shrink_to_fit=False, indent=0)
                column += 1
            row += 1
        
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
        for col, value in dims.items():
            ws.column_dimensions[col].width = value + 4 if value  < maxwidth else maxwidth

        return a

class elb(aws):
    """ELB boto3 class"""
    service = "elb"
    MaxItems = 400

    def paginator(self, call=None, key=None, **kwargs):
        a = call(**kwargs)
        base = a
        while 'NextMarker' in a:
            a = call(Marker=a['NextMarker'], **kwargs)
            base[key].extend(a[key])
        return base


class iam(aws):
    """IAM boto3 class"""
    service = "iam"
    MaxItems = 1000

    def paginator(self, call=None, key=None, **kwargs):
        a = call(**kwargs)
        base = a
        while a['IsTruncated'] is True and 'Marker' in a:
            a = call(Marker=a['Marker'], **kwargs)
            base[key].extend(a[key])
        return base

class ec2(aws):
    """EC2 boto3 class"""
    service = "ec2"
    
class AutoScalingInstances(aws):
    """AutoScalingInstances boto3 class"""
    service = "autoscaling"
    MaxRecords = 100
        
    def describe(self):
        # return self.client.describe_auto_scaling_instances()
        return self.paginator(call=self.client.describe_auto_scaling_groups, key='AutoScalingGroups', MaxRecords=self.MaxRecords)

class CloudFront(aws):
    """CloudFront boto3 class"""
    service = "cloudfront"
    
    def describe(self):
        return self.client.list_distributions()

class CloudTrails(aws):
    """CloudTrails boto3 class"""
    service = "cloudtrail"

    def describe(self):
        return self.client.describe_trails()

class DynamoDBTables(aws):
    """DynamoDBTables boto3 class"""
    service = "dynamodb"
    MaxItems = 100

    def paginator(self, call=None, key=None, **kwargs):
        a = call(**kwargs)
        base = a
        NextToken = 'LastEvaluatedTableName'
        while NextToken in a:
            a = call(ExclusiveStartTableName=a[NextToken], **kwargs)
            base[key].extend(a[key])
        return base

    def describe(self):
        key = 'TableNames'
        a = self.paginator(call=self.client.list_tables, key=key, Limit=self.MaxItems)
        tables = []
        for t in a[key]:
            table = self.client.describe_table(TableName=t)
            tables.append(table)
        return tables

class DynamoDBGlobalTables(aws):
    """DynamoDBGlobalTables boto3 class"""
    service = "dynamodb"
    MaxItems = 100

    def paginator(self, call=None, key=None, **kwargs):
        a = call(**kwargs)
        base = a
        NextToken = 'LastEvaluatedGlobalTableName'
        while NextToken in a:
            a = call(ExclusiveStartGlobalTableName=a[NextToken], **kwargs)
            base[key].extend(a[key])
        return base

    def describe(self):
        key = 'GlobalTables'
        a = self.paginator(call=self.client.list_global_tables, key=key, Limit=self.MaxItems)
        tables = []
        for t in a[key]:
            table = self.client.describe_global_table(TableName=t)
            tables.append(table)
        return tables


class EC2IAM(ec2):
    """EC2 IAM boto3 class"""

    def describe(self):
        return self.paginator(call=self.client.describe_iam_instance_profile_associations, key='IamInstanceProfileAssociations')

class EC2Instances(ec2):
    """EC2 boto3 class"""
    MaxItems = 1000

    def format_cell(self, value=None, field=None):
        (out, rc) = self.format_cell_loop(value=value, field=field)
        return out

    def append_images(self, instances=None):
        cache = {}
        for r in instances['Reservations']:
            for i in r['Instances']:
                ImageId = i['ImageId']
                if ImageId in cache:
                    image = cache[ImageId]
                else:
                    image = self.client.describe_images(ImageIds=[ImageId])
                    if type(image['Images']) is list and image['Images'] != []:
                        image = image['Images'][0]
                    else:
                        image = {
                            'Id': ImageId,
                            'Name': 'Not found',
                            'Description': ''
                        }
                    cache[ImageId] = image
      
                i['ImageName'] = image['Name']
                i['ImageDescription'] = image['Description'] if 'Description' in image else ''
        return instances

    def describe(self):
        a = self.paginator(call=self.client.describe_instances, key='Reservations', MaxResults=self.MaxItems)
        instances = self.append_images(a)
        return self.json_datetime_convert(instances)
        
class EC2AMI(ec2):
    """EC2 AMI boto3 class"""

    def describe(self):
        return self.client.describe_images(Owners=[ 'self' ])

class ElasticBlockStorages(ec2):
    """EBS boto3 class"""

    MaxItems = 500

    def describe(self):
        a = self.paginator(call=self.client.describe_volumes, key='Volumes', MaxResults=self.MaxItems)
        for i in a['Volumes']:
            if 'Attachments' not in i or i['Attachments'] == []:
                i['Attachments'] = [ {
                    'AttachTime': None,
                    'DeleteOnTermination': False,
                    'Device': None,
                    'InstanceId': None,
                    'State': 'available',
                    'VolumeId': i['VolumeId']
                } ]
        return self.json_datetime_convert(a)

class UnencryptedEBS(ElasticBlockStorages):
    """Unencrypted EBS boto3 class"""
    pass

class UnattachedEBS(ElasticBlockStorages):
    """Unattached EBS boto3 class"""
    pass
        

class ElasticIPs(ec2):
    """ElasticIPs boto3 class"""

    def describe(self):
        return self.client.describe_addresses()

class IAMGroups(iam):
    """IAM Groups boto3 class"""


    def describe(self):
        return self.paginator(call=self.client.list_groups, key='Groups', MaxItems=self.MaxItems)

class IAMPasswordPolicy(iam):
    """IAM Password Policy boto3 class"""

    def describe(self):
        return self.boto_method_handler(call=self.client.get_account_password_policy)

class IAMUsers(iam):
    """IAM Users boto3 class"""
    
    def format_cell(self, value=None, field=None):
        (out, rc) = self.format_cell_loop(value=value, field=field)
        return out

    def describe(self):
        a = self.paginator(call=self.client.list_users, key='Users', MaxItems=self.MaxItems)

        for user in a['Users']:
            n = user["UserName"]
            cmap = {
                'Groups': self.client.list_groups_for_user,
                'AttachedPolicies': self.client.list_attached_user_policies
            }
            
            print("[%s] Loading attributes for %s" % (self.service, n))
            for item in cmap:
                user[item] = self.boto_method_handler(call=cmap[item], UserName=n, MaxItems=self.MaxItems)
        return a

class IAMRoles(iam):
    """IAM Roles boto3 class"""

    def describe(self):
        return self.paginator(call=self.client.list_roles, key='Roles', MaxItems=self.MaxItems)

class IAMPolicies(iam):
    """IAM Policies boto3 class"""

    def describe(self):
        return self.paginator(call=self.client.list_policies, key='Policies', MaxItems=self.MaxItems)

class IAMAttachedGroupPolicies(iam):
    """IAM Attached Group Policies boto3 class"""

    def describe(self):
        a = self.paginator(call=self.client.list_groups, key='Groups', MaxItems=self.MaxItems)
        policies = []
        for i in a['Groups']:
            groupname = i['GroupName']
            # print("[%s] Loading attached policies for group %s" % (self.service, groupname))
            p = self.client.list_attached_group_policies(GroupName=groupname)
            p['GroupName'] = groupname
            policies.append(p)
        return policies

class LoadBalancers(elb):
    """LoadBalancers boto3 class"""
    service = "elbv2"

    def describe(self):
        a = self.paginator(call=self.client.describe_load_balancers, key='LoadBalancers', PageSize=self.MaxItems)

        for elb in a['LoadBalancers']:
            n = elb["LoadBalancerArn"]
            cmap = {
                'AccessLogEnabled': self.client.describe_load_balancer_attributes
            }

            print("[%s] Loading attributes for %s" % (self.service, n))
            for item in cmap:
                elb[item] = self.boto_method_handler(call=cmap[item], LoadBalancerArn=n)
        
        return a

class ClassicLoadBalancers(elb):
    """Classic LoadBalancers boto3 class"""
    service = "elb"

    def format_cell(self, value=None, field=None):
        return self.format_cell_yaml(value=value)

    def describe(self):
        a = self.paginator(call=self.client.describe_load_balancers, key='LoadBalancerDescriptions', PageSize=self.MaxItems)
        for elb in a['LoadBalancerDescriptions']:
            n = elb["LoadBalancerName"]
            cmap = {
                'AccessLogEnabled': self.client.describe_load_balancer_attributes
            }
            
            print("[%s] Loading attributes for %s" % (self.service, n))
            for item in cmap:
                elb[item] = self.boto_method_handler(call=cmap[item], LoadBalancerName=n)

        return a

class NATGateways(ec2):
    """NAT Gateways boto3 class"""

    def describe(self):
        return self.paginator(call=self.client.describe_nat_gateways, key='NatGateways', MaxResults=self.MaxItems)

class RDS(aws):
    """RDS boto3 class"""
    service = "rds"

    def describe(self):
        return self.paginator(call=self.client.describe_db_instances, key='DBInstances', MaxRecords=self.MaxItems)

    
class S3(aws):
    """S3 Buckets class"""
    service = "s3"

    def format_cell(self, value=None, field=None):
        (out, rc) = self.format_cell_loop(value=value, field=field)
        return out

    def describe(self):
        a = self.client.list_buckets()
        for bucket in a['Buckets']:
            n = bucket['Name']
            cmap = {
                'AccelerateConfiguration': self.client.get_bucket_accelerate_configuration,
                'EventNotifications': self.client.get_bucket_notification_configuration,
                'Versioning': self.client.get_bucket_versioning,
                'Logging': self.client.get_bucket_logging,
                'Tags': self.client.get_bucket_tagging,
                'Payer': self.client.get_bucket_request_payment,
                'ObjectLockConfiguration': self.client.get_object_lock_configuration,
                'ACL': self.client.get_bucket_acl,
                'Policy': self.client.get_bucket_policy,
                'CORS': self.client.get_bucket_cors,
                'Encryption': self.client.get_bucket_encryption,
                'Lifecycle': self.client.get_bucket_lifecycle_configuration,
                'Replication': self.client.get_bucket_replication,
                'Analytics': self.client.get_bucket_analytics_configuration,
                'Inventory': self.client.list_bucket_inventory_configurations,
                'Metrics': self.client.get_bucket_metrics_configuration,
                'Website': self.client.get_bucket_website,
                'PublicAccessBlock': self.client.get_public_access_block
            }
            
            print("[%s] Loading configuration for s3 bucket=%s" % (self.service, n))

            for item in cmap:
                bucket[item] = self.boto_method_handler(call=cmap[item], Bucket=n)

        return a
            
class SecurityGroups(ec2):
    """SecurityGroups boto3 class"""

    MaxItems = 1000

    def set_cell_color(self, value=None, field=None):
        color = None
        if value is None or field is None:
            return color
        colors = self.colors
        cfdata = self.cfdata
        field_colors = cfdata[field]['colors'] if field in cfdata and 'colors' in cfdata[field] else {}
        hash_object = hashlib.sha256(str(value))
        hex_digest = hash_object.hexdigest()
        if hex_digest in field_colors:
            color = Colors.Color(rgb=field_colors[hex_digest])
        elif value in colors:
            color = Colors.Color(rgb=colors[value])
        return color
        
    def rebase_rule_types(self):
        services = {}
        data = self.attributes['rule_types'] if 'rule_types' in self.attributes else None
        if data is None:
            return services
        for k,v in data.iteritems():
            for proto in v:
                port = v[proto]
                key = str(proto) + ':' + str(port)
                services[key] = k
        return services
          
    def format_cell(self, value=None, field=None):
        delim = ' - '
        out = ''
        services = self.rebase_rule_types()
        rules = json.loads(value)
        rc = 1
        cfdata = self.cfdata
        color = None
        colors = cfdata[field]['colors'] if field in cfdata and 'colors' in cfdata[field] else {}

        for r in rules:
            port = r['FromPort'] if 'FromPort' in r else 'All'
            toport = r['ToPort'] if 'ToPort' in r else 'All'
            proto = r['IpProtocol'] if 'IpProtocol' in r else 'All'
            if proto == '-1':
                proto = 'All'
            key = str(proto) + ':'
            if port != toport:
                port = str(port) + '-' + str(toport)
            key += str(port)
            defaultkey = str(proto) + ':' + '*'
            value = services[key] if key in services else services[defaultkey]
            h = value + delim + proto + delim + str(port) + delim
            (out, rc) = self.format_cell_loop(value=json.dumps(r), field=field, header=h, rc=rc, out=out)


        hash_object = hashlib.sha256(out)
        hex_digest = hash_object.hexdigest()

        for pattern in colors:
            if (re.search(str(pattern), out)):
                color = colors[pattern]

        if color is not None:
            colors[hex_digest] = color

        return out

    def describe(self):
        return self.paginator(call=self.client.describe_security_groups, key='SecurityGroups', MaxResults=self.MaxItems)

class VPNGateways(ec2):
    """VPNGateways boto3 class"""

    def describe(self):
        return self.client.describe_vpn_gateways()


def register_classes():
    ch = ClassHolder()
    ch.add_class(AutoScalingInstances)
    ch.add_class(CloudFront)
    ch.add_class(CloudTrails)
    ch.add_class(DynamoDBTables)
    ch.add_class(DynamoDBGlobalTables)
    ch.add_class(EC2IAM)
    ch.add_class(EC2Instances)
    ch.add_class(UnencryptedEBS)
    ch.add_class(UnattachedEBS)
    ch.add_class(EC2AMI)
    ch.add_class(ElasticBlockStorages)
    ch.add_class(ElasticIPs)
    ch.add_class(IAMGroups)
    ch.add_class(IAMUsers)
    ch.add_class(IAMPasswordPolicy)
    ch.add_class(IAMRoles)
    ch.add_class(IAMPolicies)
    ch.add_class(IAMAttachedGroupPolicies)
    ch.add_class(LoadBalancers)
    ch.add_class(ClassicLoadBalancers)
    ch.add_class(NATGateways)
    ch.add_class(RDS)
    ch.add_class(S3)
    ch.add_class(SecurityGroups)
    ch.add_class(VPNGateways)
    return ch

def main():
    wb = Workbook()
    ws = wb.active
    
    """ 
    Hightlight style for header
    https://openpyxl.readthedocs.io/en/stable/styles.html#creating-a-named-style
    """

    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True)
    bd = Side(style='thin', color="000000")
    highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    rmap = conf()
    resources = []

    ch = register_classes()

    for res in rmap['resources']:
        if type(res) is dict:
            for res_ in res:
                resources.append(ch[res_](conf=res, region=rmap['region']))

    for res in resources:
        #pp.pprint(res.describe())

        f = res.xlsx(wb, highlight)
        print("%s: %d" % (res.description, len(f)))
        #break

    wb.remove_sheet(ws)
    wb.save(rmap['of'])
        
if __name__ == '__main__':
	main()
